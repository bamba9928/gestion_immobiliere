import logging
import mimetypes
from datetime import date
from itertools import chain

import openpyxl
from django.db import transaction
from openpyxl.styles import Font, PatternFill, Alignment
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.models import Group
from django.core.exceptions import PermissionDenied
from django.core.mail import send_mail
from django.core.management import call_command
from django.db.models import Q, Sum
from django.http import FileResponse, Http404, HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views.generic import ListView, DetailView, FormView
from django.views.generic.edit import FormMixin
from django.core.exceptions import ValidationError
from .services.paiement import PaymentService
from .forms import CashPaymentForm

from .forms import (
    BienForm,
    InterventionForm,
    ContactSiteForm,
    ContactAnnonceForm,
    BailForm,
    EtatDesLieuxForm,
    LocataireCreationForm,
    DepenseForm, UnifiedCreationForm,
)
from .models import (
    Bien,
    Bail,
    Loyer,
    Intervention,
    Annonce,
    EtatDesLieux,
    Transaction,
    Depense,
)
from .permissions import (
    is_admin,
    is_bailleur,
    is_locataire,
    get_active_bail,
)
from .services.stats import DashboardService

logger = logging.getLogger(__name__)
User = get_user_model()


# ============================================================================
# PAGES PUBLIQUES
# ============================================================================

class HomeView(ListView):
    model = Annonce
    template_name = "home.html"
    context_object_name = "annonces"
    paginate_by = 9

    def get_base_queryset(self):
        return (
            Annonce.objects.filter(
                statut="PUBLIE",
                bien__in=Bien.objects.disponibles(),
            )
            .select_related("bien", "bien__proprietaire")
        )

    def get_queryset(self):
        queryset = self.get_base_queryset()

        if q := self.request.GET.get("q"):
            queryset = queryset.filter(
                Q(bien__ville__icontains=q)
                | Q(bien__adresse__icontains=q)
                | Q(titre__icontains=q)
                | Q(description__icontains=q)
            )

        if type_bien := self.request.GET.get("type"):
            queryset = queryset.filter(bien__type_bien=type_bien)

        if ville := self.request.GET.get("ville"):
            queryset = queryset.filter(bien__ville=ville)

        if prix_max := self.request.GET.get("prix_max"):
            queryset = queryset.filter(prix__lte=prix_max)

        allowed_sorts = ["-date_publication", "date_publication", "prix", "-prix"]
        sort = self.request.GET.get("sort", "-date_publication")
        if sort not in allowed_sorts:
            sort = "-date_publication"

        return queryset.order_by(sort)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        base_qs = self.get_base_queryset()
        paginator = context.get("paginator")

        context.update(
            {
                "types_bien": Bien.TYPE_CHOICES,
                "villes": (
                    base_qs.values_list("bien__ville", flat=True)
                    .distinct()
                    .order_by("bien__ville")
                ),
                "annonces_count": paginator.count if paginator else 0,
            }
        )
        return context


class AnnonceDetailView(FormMixin, DetailView):
    model = Annonce
    template_name = "biens/annonce_detail.html"
    context_object_name = "annonce"
    form_class = ContactAnnonceForm

    def get_queryset(self):
        return (
            Annonce.objects.filter(
                statut="PUBLIE",
                bien__in=Bien.objects.disponibles(),
            )
            .select_related("bien", "bien__proprietaire")
        )

    def get_success_url(self):
        return reverse("annonce_detail", kwargs={"pk": self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["annonces_similaires"] = (
            Annonce.objects.filter(
                statut="PUBLIE",
                bien__in=Bien.objects.disponibles(),
                bien__type_bien=self.object.bien.type_bien,
                bien__ville=self.object.bien.ville,
            )
            .exclude(id=self.object.id)[:3]
        )
        context.setdefault("form", self.get_form())
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.get_form()
        if form.is_valid():
            contact = form.save(commit=False)
            contact.annonce = self.object
            contact.save()
            messages.success(
                request,
                "Votre message a bien été envoyé. Nous vous recontacterons rapidement.",
            )
            return redirect(self.get_success_url())

        messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
        return self.render_to_response(self.get_context_data(form=form))


class ContactView(FormView):
    template_name = "pages/contact.html"
    form_class = ContactSiteForm
    success_url = reverse_lazy("contact")

    def form_valid(self, form):
        data = form.cleaned_data
        sujet = f"[Contact MADA IMMO] {data.get('sujet')}"
        message = (
            f"Nom : {data.get('nom')}\n"
            f"Email : {data.get('email')}\n"
            f"Téléphone : {data.get('telephone')}\n\n"
            f"Message :\n{data.get('message')}"
        )
        destinataire = getattr(settings, "CONTACT_EMAIL", "gestionadmin@votre-site.com")

        try:
            send_mail(
                sujet,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [destinataire],
                fail_silently=False,
            )
            messages.success(self.request, "Votre message a bien été envoyé, nous vous répondrons rapidement.")
        except Exception as e:
            logger.error("Erreur envoi email contact: %s", e)
            messages.error(self.request, "Une erreur est survenue. Veuillez réessayer plus tard.")

        return super().form_valid(form)


def about(request):
    return render(request, "about.html")


# ============================================================================
# COMPTABILITÉ
# ============================================================================

@login_required
def grand_livre(request):
    if not is_admin(request.user):
        raise PermissionDenied("Accès réservé aux administrateurs.")

    try:
        annee = int(request.GET.get("annee", date.today().year))
    except (TypeError, ValueError):
        annee = date.today().year

    recettes = (
        Transaction.objects.filter(est_validee=True, created_at__year=annee)
        .select_related("loyer__bail__bien", "loyer__bail__locataire")
    )
    depenses = Depense.objects.filter(date_paiement__year=annee).select_related("bien")

    total_recettes = recettes.aggregate(Sum("montant"))["montant__sum"] or 0
    total_depenses = depenses.aggregate(Sum("montant"))["montant__sum"] or 0
    cash_flow = total_recettes - total_depenses

    for r in recettes:
        r.flux = "CREDIT"
        r.date_ope = r.created_at.date()
        r.libelle_comptable = (
            f"Loyer {r.loyer.periode_debut.strftime('%m/%Y')} - {r.loyer.bail.locataire.last_name}"
        )

    for d in depenses:
        d.flux = "DEBIT"
        d.date_ope = d.date_paiement
        d.libelle_comptable = f"{d.get_type_depense_display()} : {d.libelle}"

    mouvements = sorted(chain(recettes, depenses), key=lambda x: x.date_ope, reverse=True)

    annees_recettes = Transaction.objects.filter(est_validee=True).dates("created_at", "year")
    annees_depenses = Depense.objects.all().dates("date_paiement", "year")
    annees_disponibles = sorted({d.year for d in chain(annees_recettes, annees_depenses)}, reverse=True) or [annee]

    return render(
        request,
        "comptabilite/grand_livre.html",
        {
            "mouvements": mouvements,
            "total_recettes": total_recettes,
            "total_depenses": total_depenses,
            "cash_flow": cash_flow,
            "annee_courante": annee,
            "annees_disponibles": annees_disponibles,
        },
    )


@login_required
def export_grand_livre_excel(request):
    if not is_admin(request.user):
        raise PermissionDenied("Accès réservé aux administrateurs.")

    try:
        annee = int(request.GET.get("annee", date.today().year))
    except (TypeError, ValueError):
        annee = date.today().year

    recettes = (
        Transaction.objects.filter(est_validee=True, created_at__year=annee)
        .select_related("loyer__bail__bien", "loyer__bail__locataire")
    )
    depenses = Depense.objects.filter(date_paiement__year=annee).select_related("bien")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="Grand_Livre_{annee}.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Grand Livre {annee}"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    center_align = Alignment(horizontal="center")
    currency_fmt = '#,##0 "FCFA"'

    headers = ["Date", "Type", "Libellé / Tiers", "Bien concerné", "Recette (Crédit)", "Dépense (Débit)"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Recettes
    for r in recettes:
        libelle = f"Loyer {r.loyer.periode_debut.strftime('%m/%Y')} - {r.loyer.bail.locataire.get_full_name()}"
        ws.append([r.created_at.date(), "RECETTE", libelle, r.loyer.bail.bien.titre, r.montant, 0])

    # Dépenses
    for d in depenses:
        ws.append([d.date_paiement, "DEPENSE", f"{d.get_type_depense_display()} : {d.libelle}", d.bien.titre, 0, d.montant])

    # Formats monétaires colonnes E/F
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=6):
        for cell in row:
            cell.number_format = currency_fmt

    # Largeurs colonnes
    widths = [15, 15, 48, 30, 22, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(response)
    return response


# ============================================================================
# DASHBOARD
# ============================================================================

@login_required
def dashboard(request):
    context = {"user": request.user, "today": date.today()}
    service = DashboardService()

    if is_admin(request.user):
        context["user_role"] = "ADMIN"
        context.update(service.get_admin_stats())

    elif is_bailleur(request.user):
        context["user_role"] = "BAILLEUR"
        context.update(service.get_bailleur_stats(request.user))

    elif is_locataire(request.user):
        bail = get_active_bail(request.user)

        if bail:
            loyers_qs = Loyer.objects.filter(bail=bail).order_by("-date_echeance")
            dernier_loyer = loyers_qs.first()
            interventions_qs = (
                Intervention.objects.filter(bien=bail.bien, locataire=request.user)
                .select_related("bien")
                .order_by("-created_at")[:5]
            )
        else:
            loyers_qs = Loyer.objects.none()
            dernier_loyer = None
            interventions_qs = Intervention.objects.none()

        context.update(
            {
                "user_role": "LOCATAIRE",
                "bail": bail,
                "loyers": loyers_qs,
                "dernier_loyer": dernier_loyer,
                "interventions_recents": interventions_qs,
                "has_active_bail": bool(bail),
            }
        )
    else:
        context["user_role"] = "AGENT"

    return render(request, "dashboard.html", context)


# ============================================================================
# BIENS
# ============================================================================

@login_required
def add_bien(request):
    if not (is_admin(request.user) or is_bailleur(request.user)):
        raise PermissionDenied("Seuls les administrateurs et propriétaires peuvent ajouter des biens.")

    if request.method == "POST":
        form = BienForm(request.POST, request.FILES)
        if form.is_valid():
            bien = form.save(commit=False)
            bien.proprietaire = request.user
            bien.save()
            messages.success(request, f"Le bien '{bien.titre}' a été ajouté avec succès !")
            return redirect("dashboard")
    else:
        form = BienForm()

    return render(
        request,
        "biens/add_bien.html",
        {"form": form, "user_role": "ADMIN" if is_admin(request.user) else "BAILLEUR"},
    )


@login_required
def gestion_bien_detail(request, pk):
    bien = get_object_or_404(Bien, pk=pk)

    if not (is_admin(request.user) or bien.proprietaire == request.user):
        raise PermissionDenied("Vous n'êtes pas propriétaire de ce bien.")

    # Bail actif = date_fin NULL ou date_fin >= aujourd’hui
    bail_actif = (
        bien.baux.filter(
            est_signe=True,
            date_debut__lte=date.today(),
            date_fin__gte=date.today(),
        )
        .order_by("-date_debut")
        .first()
    )

    interventions = bien.interventions.select_related("locataire").order_by("-created_at")[:5]

    total_recettes = (
        Transaction.objects.filter(loyer__bail__bien=bien, est_validee=True)
        .aggregate(Sum("montant"))["montant__sum"] or 0
    )
    total_depenses = Depense.objects.filter(bien=bien).aggregate(Sum("montant"))["montant__sum"] or 0
    cash_flow = total_recettes - total_depenses

    return render(
        request,
        "biens/gestion_detail.html",
        {
            "bien": bien,
            "bail_actif": bail_actif,
            "interventions": interventions,
            "total_recettes": total_recettes,
            "total_depenses": total_depenses,
            "cash_flow": cash_flow,
            "est_loue": bail_actif is not None,
        },
    )


@login_required
def biens_list(request):
    if is_admin(request.user):
        qs = Bien.objects.select_related("proprietaire").order_by("-created_at")
        user_role = "ADMIN"
    elif is_bailleur(request.user):
        qs = Bien.objects.filter(proprietaire=request.user).order_by("-created_at")
        user_role = "BAILLEUR"
    else:
        raise PermissionDenied("Accès réservé.")

    return render(request, "biens/biens_list.html", {"biens": qs, "user_role": user_role})


@login_required
def edit_bien(request, pk):
    bien = get_object_or_404(Bien, pk=pk)

    if not (is_admin(request.user) or bien.proprietaire == request.user):
        raise PermissionDenied("Accès refusé.")

    if request.method == "POST":
        form = BienForm(request.POST, request.FILES, instance=bien)
        if form.is_valid():
            form.save()
            messages.success(request, "Bien mis à jour.")
            return redirect("gestion_bien_detail", pk=bien.pk)
    else:
        form = BienForm(instance=bien)

    return render(request, "biens/edit_bien.html", {"form": form, "bien": bien})


# ============================================================================
# BAUX
# ============================================================================

@login_required
def add_bail(request):
    if not (is_admin(request.user) or is_bailleur(request.user)):
        raise PermissionDenied("Seuls les administrateurs et bailleurs peuvent créer un bail.")

    biens_queryset = Bien.objects.filter(proprietaire=request.user) if is_bailleur(request.user) else Bien.objects.all()

    if request.method == "POST":
        form = BailForm(request.POST, request.FILES)
        form.fields["bien"].queryset = biens_queryset
        if form.is_valid():
            bail = form.save()
            try:
                from apps.core.services.contrat import sauvegarder_contrat
                sauvegarder_contrat(bail)
                messages.success(request, "Le bail a été créé avec succès et le contrat PDF généré.")
            except Exception as e:
                logger.error("Erreur génération contrat PDF bail %s: %s", bail.id, e)
                messages.warning(request, "Le bail a été créé, mais la génération du PDF a échoué.")
            return redirect("bail_detail", pk=bail.pk)
    else:
        form = BailForm()
        form.fields["bien"].queryset = biens_queryset

    return render(
        request,
        "baux/add_bail.html",
        {"form": form, "user_role": "ADMIN" if is_admin(request.user) else "BAILLEUR"},
    )


@login_required
def bail_detail(request, pk):
    bail = get_object_or_404(Bail, pk=pk)

    if not (is_admin(request.user) or bail.bien.proprietaire == request.user or bail.locataire == request.user):
        raise PermissionDenied("Vous n'avez pas l'autorisation de consulter ce bail.")

    loyers = bail.loyers.order_by("-date_echeance")
    etats = bail.etats_des_lieux.order_by("-date_realisation")
    edl_entree = etats.filter(type_edl="ENTREE").first()
    edl_sortie = etats.filter(type_edl="SORTIE").first()

    if is_admin(request.user):
        user_role = "ADMIN"
    elif bail.bien.proprietaire == request.user:
        user_role = "BAILLEUR"
    else:
        user_role = "LOCATAIRE"

    return render(
        request,
        "baux/bail_detail.html",
        {
            "bail": bail,
            "loyers": loyers,
            "etats": etats,
            "edl_entree": edl_entree,
            "edl_sortie": edl_sortie,
            "user_role": user_role,
        },
    )


@login_required
def add_etat_des_lieux(request, bail_id, type_edl):
    bail = get_object_or_404(Bail, pk=bail_id)

    if not (is_admin(request.user) or bail.bien.proprietaire == request.user or bail.locataire == request.user):
        raise PermissionDenied("Vous n'avez pas l'autorisation de créer cet état des lieux.")

    if type_edl not in ["ENTREE", "SORTIE"]:
        messages.error(request, "Type d'état des lieux invalide.")
        return redirect("bail_detail", pk=bail.pk)

    instance = EtatDesLieux(bail=bail, type_edl=type_edl, date_realisation=timezone.now().date())
    disabled_fields = ["bail", "type_edl"]

    if request.method == "POST":
        form = EtatDesLieuxForm(request.POST, request.FILES, instance=instance)
        for f in disabled_fields:
            if f in form.fields:
                form.fields[f].disabled = True
        if form.is_valid():
            form.save()
            messages.success(request, "L'état des lieux a été enregistré avec succès.")
            return redirect("bail_detail", pk=bail.pk)
    else:
        form = EtatDesLieuxForm(instance=instance)
        for f in disabled_fields:
            if f in form.fields:
                form.fields[f].disabled = True

    return render(request, "etats_des_lieux/form.html", {"form": form, "bail": bail, "type_edl": type_edl})


@login_required
def generate_lease_pdf(request, bail_id):
    bail = get_object_or_404(Bail, pk=bail_id)

    if not (is_admin(request.user) or bail.bien.proprietaire == request.user):
        raise PermissionDenied("Vous n'avez pas le droit de générer ce contrat.")

    try:
        from apps.core.services.contrat import sauvegarder_contrat
        sauvegarder_contrat(bail)
        messages.success(request, "Contrat PDF généré et attaché au bail.")
    except Exception as e:
        logger.error("Erreur génération contrat PDF bail %s: %s", bail.id, e)
        messages.error(request, "Impossible de générer le contrat. Contactez l'administrateur.")

    return redirect("bail_detail", pk=bail.pk)


@login_required
def download_contrat(request, bail_id):
    bail = get_object_or_404(Bail, pk=bail_id)

    is_proprio = bail.bien.proprietaire == request.user
    is_loc = bail.locataire == request.user

    if not (is_admin(request.user) or is_proprio or is_loc):
        raise PermissionDenied("Vous n'avez pas l'autorisation de télécharger ce contrat.")

    if not bail.fichier_contrat:
        raise Http404("Aucun contrat signé n'est disponible pour ce bail.")

    response = FileResponse(bail.fichier_contrat.open("rb"), content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="Bail_{bail.id}.pdf"'
    return response


# ============================================================================
# LOYERS / PAIEMENTS
# ============================================================================

@login_required
def loyers_list(request):
    if not is_admin(request.user):
        raise PermissionDenied("Accès réservé aux administrateurs.")

    statut = request.GET.get("statut")
    base_qs = Loyer.objects.select_related("bail__locataire", "bail__bien")
    loyers = base_qs.order_by("-date_echeance")
    if statut:
        loyers = loyers.filter(statut=statut)

    stats = {
        "total_retard": base_qs.filter(statut="RETARD").count(),
        "total_attente": base_qs.filter(statut="A_PAYER").count(),
    }

    return render(request, "interventions/loyers_list.html", {"loyers": loyers, "current_statut": statut, "stats": stats})


@login_required
def admin_process_cash_payment(request, loyer_id):
    """
    Encaissement manuel (CASH) par l'admin.
    Crée une transaction comptable et met à jour le loyer.
    """
    if not is_admin(request.user):
        raise PermissionDenied("Seul un administrateur peut encaisser des espèces.")

    loyer = get_object_or_404(Loyer, id=loyer_id)

    if loyer.statut == "PAYE":
        messages.info(request, "Ce loyer est déjà intégralement payé.")
        return redirect("loyers_list")

    if request.method == "POST":
        form = CashPaymentForm(request.POST, loyer=loyer)
        if form.is_valid():
            montant = form.cleaned_data['montant']
            try:
                service = PaymentService()
                service.enregistrer_paiement_especes(
                    loyer=loyer,
                    montant=montant,
                    auteur_admin=request.user
                )

                messages.success(request, f"Paiement de {montant} FCFA enregistré (Espèces).")

                if loyer.statut == "PAYE":
                    messages.success(request, "Le loyer est soldé. La quittance a été générée.")

                return redirect("loyers_list")

            except Exception as e:
                logger.error(f"Erreur paiement cash : {e}")
                messages.error(request, f"Erreur : {e}")
    else:
        form = CashPaymentForm(loyer=loyer)

    return render(request, "paiements/admin_cash_form.html", {
        "form": form,
        "loyer": loyer
    })
@login_required
def download_quittance(request, loyer_id):
    loyer = get_object_or_404(
        Loyer.objects.select_related("bail__locataire", "bail__bien__proprietaire"),
        id=loyer_id
    )

    is_owner = (loyer.bail.bien.proprietaire == request.user)

    if not (is_admin(request.user) or is_owner or loyer.bail.locataire == request.user):
        raise PermissionDenied("Accès non autorisé à cette quittance.")

    if loyer.statut != "PAYE":
        messages.error(request, "La quittance n'est disponible que pour les loyers payés.")
        return redirect("dashboard")

    if not loyer.quittance:
        try:
            from apps.core.services.quittance import attacher_quittance
            attacher_quittance(loyer)
        except Exception as e:
            logger.error("Erreur génération quittance loyer %s: %s", loyer.id, e)
            messages.error(request, "Impossible de générer la quittance. Contactez l'administrateur.")
            return redirect("dashboard")

    if not loyer.quittance:
        messages.error(request, "La quittance n'est pas disponible pour le moment.")
        return redirect("dashboard")

    return FileResponse(
        loyer.quittance.open("rb"),
        content_type="application/pdf",
        filename=loyer.quittance.name.split("/")[-1],
        as_attachment=False,
    )


# ============================================================================
# INTERVENTIONS
# ============================================================================

@login_required
def interventions_list(request):
    user_is_admin = is_admin(request.user)
    user_is_bailleur = is_bailleur(request.user)

    bail = None
    if not user_is_admin and not user_is_bailleur:
        bail = get_active_bail(request.user)

    if user_is_admin:
        interventions = Intervention.objects.select_related("bien", "locataire", "bien__proprietaire").order_by("-created_at")
    elif user_is_bailleur:
        interventions = (
            Intervention.objects.filter(bien__proprietaire=request.user)
            .select_related("bien", "locataire")
            .order_by("-created_at")
        )
    elif bail:
        interventions = Intervention.objects.filter(bien=bail.bien, locataire=request.user).select_related("bien").order_by("-created_at")
    else:
        return render(request, "interventions/pas_de_bail.html", {"message": "Vous n'avez aucun bail actif."})

    form = InterventionForm()

    if request.method == "POST":
        if user_is_admin or user_is_bailleur:
            messages.error(request, "Seuls les locataires peuvent créer une demande d'intervention ici.")
            return redirect("interventions_list")
        if not bail:
            messages.error(request, "Impossible de créer une intervention sans bail actif.")
            return redirect("interventions_list")

        form = InterventionForm(request.POST, request.FILES)
        if form.is_valid():
            intervention = form.save(commit=False)
            intervention.locataire = request.user
            intervention.bien = bail.bien
            intervention.save()
            messages.success(request, "Demande d'intervention enregistrée avec succès.")
            return redirect("interventions_list")

    role_str = "ADMIN" if user_is_admin else ("BAILLEUR" if user_is_bailleur else "LOCATAIRE")

    return render(
        request,
        "interventions/interventions_list.html",
        {"interventions": interventions, "form": form, "user_role": role_str, "bail": bail},
    )


# ============================================================================
# ACTIONS ADMIN
# ============================================================================
import threading
from django.core.management import call_command


# ... vos autres imports

@login_required
def trigger_rent_generation(request):
    if not is_admin(request.user):
        raise PermissionDenied("Seul un administrateur peut générer les loyers.")

    if request.method != "POST":
        messages.warning(request, "Méthode non autorisée.")
        return redirect("dashboard")

    # Fonction locale pour exécuter la commande dans un thread séparé
    def run_command():
        try:
            call_command("generer_loyers")
            logger.info("Génération asynchrone des loyers terminée avec succès.")
        except Exception as e:
            logger.error(f"Erreur lors de la génération asynchrone : {e}")

    # Lancement du thread
    try:
        thread = threading.Thread(target=run_command)
        thread.start()

        messages.info(request,
                      "⚙️ La génération des loyers a démarré en arrière-plan. Cela peut prendre quelques instants.")
    except Exception as e:
        logger.error(f"Impossible de lancer le thread : {e}")
        messages.error(request, "Une erreur est survenue lors du lancement de l'opération.")

    return redirect("dashboard")
# ============================================================================
# UTILISATEURS
# ============================================================================

@login_required
def add_locataire(request):
    if not is_admin(request.user):
        raise PermissionDenied("Accès réservé.")

    if request.method == "POST":
        form = LocataireCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            group, _ = Group.objects.get_or_create(name="LOCATAIRE")
            user.groups.add(group)
            messages.success(request, f"Locataire {user.first_name} créé avec succès (Téléphone : {user.profile.telephone}).")
            return redirect("dashboard")
    else:
        form = LocataireCreationForm()

    return render(request, "utilisateurs/add_locataire.html", {"form": form, "user_role": "ADMIN"})


@login_required
def add_bailleur(request):
    if not is_admin(request.user):
        raise PermissionDenied("Accès réservé.")

    if request.method == "POST":
        form = LocataireCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            group, _ = Group.objects.get_or_create(name="BAILLEUR")
            user.groups.add(group)
            messages.success(request, f"Bailleur {user.get_full_name() or user.username} créé avec succès.")
            return redirect("dashboard")
    else:
        form = LocataireCreationForm()

    return render(request, "utilisateurs/add_bailleur.html", {"form": form, "user_role": "ADMIN"})


@login_required
def add_agent(request):
    if not is_admin(request.user):
        raise PermissionDenied("Accès réservé.")

    if request.method == "POST":
        form = LocataireCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            group, _ = Group.objects.get_or_create(name="AGENT")
            user.groups.add(group)
            messages.success(request, f"Agent {user.get_full_name() or user.username} créé avec succès.")
            return redirect("dashboard")
    else:
        form = LocataireCreationForm()

    return render(request, "utilisateurs/add_agent.html", {"form": form, "user_role": "ADMIN"})


# ============================================================================
# DÉPENSES / GED / KYC
# ============================================================================

@login_required
def add_depense(request):
    if not (is_admin(request.user) or is_bailleur(request.user)):
        raise PermissionDenied("Vous n'avez pas les droits pour ajouter une dépense.")

    if request.method == "POST":
        form = DepenseForm(request.POST, request.FILES)
        if is_bailleur(request.user):
            form.fields["bien"].queryset = Bien.objects.filter(proprietaire=request.user)

        if form.is_valid():
            depense = form.save()
            messages.success(request, f"Dépense '{depense.libelle}' enregistrée avec succès.")
            return redirect("grand_livre" if is_admin(request.user) else "dashboard")

        messages.error(request, "Erreur dans le formulaire. Veuillez vérifier les champs.")
    else:
        form = DepenseForm()
        if is_bailleur(request.user):
            form.fields["bien"].queryset = Bien.objects.filter(proprietaire=request.user)

    return render(
        request,
        "comptabilite/add_depense.html",
        {"form": form, "user_role": "ADMIN" if is_admin(request.user) else "BAILLEUR"},
    )


@login_required
def documents_list(request):
    user = request.user

    baux = Bail.objects.none()
    quittances = Loyer.objects.none()
    edls = EtatDesLieux.objects.none()

    if is_admin(user):
        baux = Bail.objects.select_related("bien", "locataire").filter(fichier_contrat__isnull=False).exclude(fichier_contrat="")
        quittances = Loyer.objects.select_related("bail__locataire", "bail__bien").filter(quittance__isnull=False).exclude(quittance="")
        edls = EtatDesLieux.objects.select_related("bail__bien").filter(pdf__isnull=False).exclude(pdf="")

    elif is_bailleur(user):
        biens_ids = Bien.objects.filter(proprietaire=user).values_list("id", flat=True)
        baux = Bail.objects.filter(bien_id__in=biens_ids).select_related("locataire", "bien").filter(fichier_contrat__isnull=False).exclude(fichier_contrat="")
        quittances = Loyer.objects.filter(bail__bien_id__in=biens_ids).select_related("bail__locataire", "bail__bien").filter(quittance__isnull=False).exclude(quittance="")
        edls = EtatDesLieux.objects.filter(bail__bien_id__in=biens_ids).select_related("bail__bien").filter(pdf__isnull=False).exclude(pdf="")

    elif is_locataire(user):
        baux = Bail.objects.filter(locataire=user).select_related("bien").filter(fichier_contrat__isnull=False).exclude(fichier_contrat="")
        quittances = Loyer.objects.filter(bail__locataire=user).select_related("bail__bien").filter(quittance__isnull=False).exclude(quittance="")
        edls = EtatDesLieux.objects.filter(bail__locataire=user).select_related("bail__bien").filter(pdf__isnull=False).exclude(pdf="")

    return render(request, "documents/ged_list.html", {"baux": baux, "quittances": quittances, "edls": edls})


@login_required
def download_kyc(request, user_id, doc_type):
    target_user = get_object_or_404(User, pk=user_id)

    if not (is_admin(request.user) or request.user == target_user):
        raise PermissionDenied("Accès refusé aux documents personnels.")

    profile = getattr(target_user, "profile", None)
    if not profile:
        raise Http404("Profil utilisateur introuvable.")

    file_obj = None
    if doc_type == "cni":
        file_obj = profile.cni_scan
    elif doc_type == "justificatif":
        file_obj = profile.justificatif_domicile

    if not file_obj:
        raise Http404("Document non trouvé.")

    content_type, _ = mimetypes.guess_type(file_obj.name)
    content_type = content_type or "application/octet-stream"

    return FileResponse(file_obj.open("rb"), content_type=content_type, filename=file_obj.name.split("/")[-1])


# ============================================================================
# BACKOFFICE LISTES (ADMIN)
# ============================================================================

class AdminRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    def test_func(self):
        return is_admin(self.request.user)


class AdminBienListView(AdminRequiredMixin, ListView):
    model = Bien
    template_name = "pages/liste_biens.html"
    context_object_name = "biens"
    paginate_by = 20
    ordering = ["-created_at"]

    def get_queryset(self):
        qs = Bien.objects.select_related("proprietaire").all()
        q = self.request.GET.get("q")
        if q:
            qs = qs.filter(Q(titre__icontains=q) | Q(ville__icontains=q) | Q(proprietaire__username__icontains=q))
        return qs
class AdminBailListView(AdminRequiredMixin, ListView):
    model = Bail
    template_name = "pages/liste_baux.html"
    context_object_name = "baux"
    paginate_by = 20
    ordering = ["-date_debut"]

    def get_queryset(self):
        qs = Bail.objects.select_related("bien", "locataire").all()
        statut = self.request.GET.get("statut")
        if statut == "actif":
            qs = qs.filter(
                est_signe=True,
                date_debut__lte=date.today(),
                date_fin__gte=date.today()
            )
        return qs
class AdminLocataireListView(AdminRequiredMixin, ListView):
    model = User
    template_name = "pages/liste_locataires.html"
    context_object_name = "locataires"
    paginate_by = 20

    def get_queryset(self):
        return (
            User.objects.filter(groups__name="LOCATAIRE")
            .select_related("profile")
            .distinct()
            .order_by("last_name")
        )


class AdminBailleurListView(AdminRequiredMixin, ListView):
    model = User
    template_name = "pages/liste_bailleurs.html"
    context_object_name = "bailleurs"
    paginate_by = 20

    def get_queryset(self):
        return (
            User.objects.filter(groups__name="BAILLEUR")
            .select_related("profile")
            .distinct()
            .order_by("last_name")
        )
@login_required
@transaction.atomic
def unified_creation_view(request):
    if not (is_admin(request.user) or is_bailleur(request.user)):
        raise PermissionDenied("Accès réservé.")

    if request.method == "POST":
        # Note: Passer l'utilisateur au formulaire pour le champ 'proprietaire'
        form = UnifiedCreationForm(request.POST, request.FILES, user=request.user)

        if form.is_valid():
            cd = form.cleaned_data
            email = cd["email"]
            user_password = cd.get("password")

            try:
                # 1) Gestion du Locataire
                locataire = User.objects.filter(Q(email__iexact=email) | Q(username__iexact=email)).first()
                created_locataire = False

                if not locataire:
                    # Génération si vide (sécurité supplémentaire)
                    if not user_password:
                        import secrets, string
                        user_password = "".join(secrets.choice(string.ascii_letters + string.digits) for _ in range(12))

                    locataire = User.objects.create_user(
                        username=email,
                        email=email,
                        first_name=cd["first_name"],
                        last_name=cd["last_name"],
                        password=user_password,
                    )
                    created_locataire = True

                # Attribution du groupe
                g_loc, _ = Group.objects.get_or_create(name="LOCATAIRE")
                locataire.groups.add(g_loc)

                # Mise à jour du profil (Correction des noms de champs ici)
                profile = getattr(locataire, "profile", None)
                if profile:
                    profile.telephone = cd.get("telephone") # Correction de phone_number
                    profile.cni_numero = cd.get("cni_numero")
                    profile.save()

                # 2) Détermination du Propriétaire
                proprietaire = cd.get("proprietaire") if (is_admin(request.user) and cd.get("proprietaire")) else request.user

                # 3) Bien Immobilier
                bien = Bien.objects.create(
                    proprietaire=proprietaire,
                    titre=cd["titre_bien"],
                    type_bien=cd["type_bien"],
                    adresse=cd["adresse"],
                    ville=cd.get("ville", "Dakar"),
                    surface=cd["surface"],
                    nb_pieces=cd.get("nb_pieces") or 1,
                    loyer_ref=cd["montant_loyer"],
                    charges_ref=cd.get("montant_charges") or 0,
                    est_actif=True,
                    description=cd.get("description", ""),
                )

                # 4) Création du Bail
                bail = Bail.objects.create(
                    bien=bien,
                    locataire=locataire,
                    date_debut=cd["date_debut"],
                    date_fin=cd["date_fin"],
                    montant_loyer=cd["montant_loyer"],
                    montant_charges=cd.get("montant_charges") or 0,
                    depot_garantie=cd["depot_garantie"],
                    jour_paiement=cd.get("jour_paiement") or 5,
                    est_signe=True,
                )

                # Génération du contrat PDF
                try:
                    from apps.core.services.contrat import sauvegarder_contrat
                    sauvegarder_contrat(bail)
                except Exception as e:
                    logger.error(f"Erreur PDF : {e}")

                if created_locataire:
                    messages.success(request, f"Locataire créé. ID : {locataire.email} | Pass : {user_password}")
                else:
                    messages.success(request, "Bail créé avec succès pour un locataire existant.")

                return redirect("bail_detail", pk=bail.pk)

            except Exception as e:
                logger.error(f"Erreur Unified View : {e}")
                messages.error(request, f"Une erreur est survenue : {e}")
    else:
        # Important : passer l'user ici aussi pour le champ proprietaire
        form = UnifiedCreationForm(user=request.user)

    return render(request, "utilisateurs/add_unified.html", {"form": form})
@login_required
def telecharger_contrat_bail(request, bail_id):
    # On récupère le bail en vérifiant que l'utilisateur est soit le locataire, soit le proprio, soit admin
    bail = get_object_or_404(Bail, id=bail_id)

    is_owner = (bail.bien.proprietaire == request.user)
    is_tenant = (bail.locataire == request.user)

    if not (request.user.is_staff or is_owner or is_tenant):
        raise PermissionDenied("Vous n'avez pas l'autorisation de consulter ce document.")

    if not bail.fichier_contrat:
        messages.error(request, "Le fichier du contrat n'est pas encore généré.")
        return redirect('dashboard')

    # Ouverture du fichier en mode binaire
    try:
        response = FileResponse(
            bail.fichier_contrat.open('rb'),
            content_type='application/pdf'
        )
        # 'inline' pour ouvrir dans le navigateur, 'attachment' pour forcer le téléchargement
        response['Content-Disposition'] = f'inline; filename="Contrat_Bail_MADA_{bail.id}.pdf"'
        return response
    except FileNotFoundError:
        raise Http404("Le fichier physique est introuvable sur le serveur.")